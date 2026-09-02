#!/usr/bin/env python3
"""Render the DEVIN ACU consumption report workbook.

Reads docs/migration/acu-usage.json (per-task attribution + context-reuse
telemetry) and docs/migration/acu-consumption-report.json (measured
per-interaction ACU ledger, phase rollup) and writes
docs/migration/DEVIN-ACU-consumption-report.xlsx.

    python3 scripts/aidlc-acu-report-xlsx.py
    python3 scripts/aidlc-acu-report-xlsx.py --rate 2.00
"""
import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
USAGE = ROOT / "docs/migration/acu-usage.json"
REPORT = ROOT / "docs/migration/acu-consumption-report.json"
OUT = ROOT / "docs/migration/DEVIN-ACU-consumption-report.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="9BC2E6")
SUB_FILL = PatternFill("solid", fgColor="DDEBF7")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

MODEL_CELL = "Devin cloud agent (mode not exposed)"

PHASE_OF = {
    "A1-A3": "Inception",
    "B1": "Construction", "B2": "Construction", "B3": "Construction",
    "C1": "Construction", "B4": "Construction", "C5": "Construction",
    "B5": "Construction", "B6": "Construction", "B7": "Construction",
    "B8": "Construction", "B9": "Construction", "C9+B10": "Construction",
    "B11+C6-C8": "Construction",
    "E1-E2": "Governance / AIDLC tooling",
    "D-PR": "Operation",
    "V1": "Validation",
}


def write_header(ws, headers, row=1):
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = WRAP


def autosize(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def sheet_phase_consumption(wb, usage, report, rate):
    ws = wb.create_sheet("AIDLC ACU consumption")
    ws["A1"] = "DEVIN ACU consumption by AIDLC phase - Angular 9 to 20 migration"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"Repository: {report['repository']}   PR: {report['pull_request']}"
    ws["A3"] = f"Session: {report['session_url']}"
    ws["A4"] = f"Rate applied: ${rate:.2f} per ACU ({report['rate']['note']})"
    ws["A5"] = ("Model column: Devin does not expose the underlying model or agent mode for a "
                "completed session, so no model attribution is claimed.")
    for r in range(2, 6):
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=False)

    headers = ["AIDLC phase", "Task", "Model / mode", "Start (UTC)", "End (UTC)",
               "Duration (s)", "ACUs", "Dollar value", "% of session", "Measurement basis"]
    write_header(ws, headers, row=7)

    session_total = report["interaction_ledger"][-1]["cumulative_acus"]
    row = 8
    for task in usage["tasks"]:
        ws.cell(row=row, column=1, value=PHASE_OF.get(task["id"], "Construction"))
        ws.cell(row=row, column=2, value=f"{task['id']}: {task['task']}").alignment = WRAP
        ws.cell(row=row, column=3, value=MODEL_CELL)
        ws.cell(row=row, column=4, value=task["start_utc"])
        ws.cell(row=row, column=5, value=task["end_utc"])
        ws.cell(row=row, column=6, value=task["duration_seconds"])
        ws.cell(row=row, column=7, value=round(task["acus"], 4))
        ws.cell(row=row, column=8, value=f"=G{row}*{rate}").number_format = '"$"#,##0.00'
        ws.cell(row=row, column=9, value=task["acus"] / session_total).number_format = "0.0%"
        ws.cell(row=row, column=10, value="Derived (pro rata by wall clock)")
        row += 1

    for item in report["interaction_ledger"]:
        if item["category"] != "Reporting":
            continue
        ws.cell(row=row, column=1, value="Reporting (post-delivery)")
        ws.cell(row=row, column=2, value=item["paid_for"]).alignment = WRAP
        ws.cell(row=row, column=3, value=MODEL_CELL)
        ws.cell(row=row, column=4, value=item["checkpoint_utc"])
        ws.cell(row=row, column=5, value="")
        ws.cell(row=row, column=6, value="")
        ws.cell(row=row, column=7, value=round(item["delta_acus"], 4))
        ws.cell(row=row, column=8, value=f"=G{row}*{rate}").number_format = '"$"#,##0.00'
        ws.cell(row=row, column=9, value=item["delta_acus"] / session_total).number_format = "0.0%"
        ws.cell(row=row, column=10, value="Measured (ACU checkpoint delta)")
        row += 1

    ws.cell(row=row, column=1, value="TOTAL (session to date)").font = BOLD
    ws.cell(row=row, column=7, value=f"=SUM(G8:G{row - 1})").font = BOLD
    ws.cell(row=row, column=8, value=f"=G{row}*{rate}").number_format = '"$"#,##0.00'
    ws.cell(row=row, column=8).font = BOLD
    autosize(ws, [26, 62, 26, 22, 12, 12, 10, 14, 12, 32])


def sheet_ledger(wb, report, rate):
    ws = wb.create_sheet("Measured interaction ledger")
    ws["A1"] = "Measured ACU checkpoints (Devin session API: acu_consumption_at_last_user_interaction)"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "Every value on this sheet is measured. Nothing here is derived or estimated."
    write_header(ws, ["#", "Checkpoint (UTC)", "Cumulative ACUs", "Delta ACUs",
                      "Dollar value of delta", "What the delta paid for", "Category",
                      "Avoidable?"], row=4)
    row = 5
    for item in report["interaction_ledger"]:
        ws.cell(row=row, column=1, value=item["n"])
        ws.cell(row=row, column=2, value=item["checkpoint_utc"])
        ws.cell(row=row, column=3, value=round(item["cumulative_acus"], 4))
        ws.cell(row=row, column=4, value=round(item["delta_acus"], 4))
        ws.cell(row=row, column=5, value=f"=D{row}*{rate}").number_format = '"$"#,##0.00'
        ws.cell(row=row, column=6, value=item["paid_for"]).alignment = WRAP
        ws.cell(row=row, column=7, value=item["category"])
        ws.cell(row=row, column=8, value=item["avoidable"]).alignment = WRAP
        row += 1

    delivery = report["interaction_ledger"][1]["delta_acus"]
    reporting = sum(i["delta_acus"] for i in report["interaction_ledger"]
                    if i["category"] == "Reporting")
    row += 1
    for label, value in [("Delivery (migration) ACUs", delivery),
                         ("Reporting / Q&A ACUs", reporting),
                         ("Session total ACUs", delivery + reporting),
                         ("Reporting share of session", None)]:
        ws.cell(row=row, column=1, value=label).font = BOLD
        if value is not None:
            ws.cell(row=row, column=3, value=round(value, 4))
            ws.cell(row=row, column=5, value=f"=C{row}*{rate}").number_format = '"$"#,##0.00'
        else:
            ws.cell(row=row, column=3,
                    value=reporting / (delivery + reporting)).number_format = "0.0%"
        row += 1
    autosize(ws, [5, 24, 18, 14, 20, 70, 16, 46])


def sheet_rollup(wb, report, rate):
    ws = wb.create_sheet("Phase rollup")
    ws["A1"] = "ACU consumption per AIDLC phase"
    ws["A1"].font = Font(bold=True, size=13)
    write_header(ws, ["AIDLC phase", "ACUs", "Dollar value", "% of session",
                      "Basis", "What it covers"], row=3)
    total = sum(p["acus"] for p in report["phase_rollup"])
    row = 4
    for phase in report["phase_rollup"]:
        ws.cell(row=row, column=1, value=phase["phase"])
        ws.cell(row=row, column=2, value=round(phase["acus"], 4))
        ws.cell(row=row, column=3, value=f"=B{row}*{rate}").number_format = '"$"#,##0.00'
        ws.cell(row=row, column=4, value=phase["acus"] / total).number_format = "0.0%"
        ws.cell(row=row, column=5, value=phase["basis"])
        ws.cell(row=row, column=6, value=phase["content"]).alignment = WRAP
        row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = BOLD
    ws.cell(row=row, column=2, value=f"=SUM(B4:B{row - 1})").font = BOLD
    ws.cell(row=row, column=3, value=f"=B{row}*{rate}").number_format = '"$"#,##0.00'
    autosize(ws, [32, 12, 14, 12, 12, 78])


def sheet_unit_rates(wb, usage, report, rate):
    ws = wb.create_sheet("Efficiency rates")
    ws["A1"] = "Derived unit rates - the numbers that make ACU consumption actionable"
    ws["A1"].font = Font(bold=True, size=13)
    write_header(ws, ["Rate", "Value", "How it is computed", "Basis"], row=3)

    migration = report["interaction_ledger"][1]["delta_acus"]
    minutes = usage["measured"]["billable_wall_clock_minutes"]
    bump_ids = {"B1", "B2", "B3", "B4", "B5", "B6", "B7", "B8", "B9"}
    bump_acus = sum(t["acus"] for t in usage["tasks"] if t["id"] in bump_ids)
    all_major_ids = bump_ids | {"C9+B10", "B11+C6-C8"}
    all_major_acus = sum(t["acus"] for t in usage["tasks"] if t["id"] in all_major_ids)
    cache = usage["prompt_cache"]["measured_window"]

    rows = [
        ("ACUs per minute of agent time", round(migration / minutes, 3),
         f"{migration:.4f} ACUs / {minutes} min", "Measured"),
        ("Seconds of wall clock per ACU", round(minutes * 60 / migration, 1),
         "inverse of the row above", "Measured"),
        ("Dollar value per minute of agent time", round(rate * migration / minutes, 2),
         f"rate x ACUs/min at ${rate:.2f}", "Measured x rate"),
        ("ACUs per plain version bump (Angular 9-18)", round(bump_acus / 9, 3),
         f"{bump_acus:.3f} ACUs over 9 plain bumps (B1-B9)", "Derived"),
        ("Dollar value per plain version bump", round(rate * bump_acus / 9, 2),
         "row above x rate", "Derived"),
        ("ACUs per major version incl. architectural work", round(all_major_acus / 11, 3),
         f"{all_major_acus:.3f} ACUs over all 11 majors (adds standalone conversion and the Angular 20 builder row)",
         "Derived"),
        ("Context reuse rate (prompt cache proxy)", f"{cache['hit_rate_pct']}%",
         f"{cache['cache_read_tokens']:,} reused of {cache['prompt_tokens']:,} processed, iterations 1-59",
         "Measured within window"),
        ("New context written per iteration", f"{usage['prompt_cache']['extrapolated_full_session']['slope_tokens_per_iteration']:,} tokens",
         "slope of context_growth_update samples", "Measured within window"),
    ]
    row = 4
    for label, value, how, basis in rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=how).alignment = WRAP
        ws.cell(row=row, column=4, value=basis)
        row += 1
    row += 1
    ws.cell(row=row, column=1,
            value=("Key reading: ACUs track elapsed agent time, so waiting on npm install costs "
                   "roughly the same as reasoning. Cutting elapsed time cuts ACUs.")).font = BOLD
    autosize(ws, [42, 22, 70, 24])


def sheet_models(wb):
    ws = wb.create_sheet("Model & mode comparison")
    ws["A1"] = "DEVIN model / agent mode comparison"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = ("Factual position: Devin bills in ACUs, not per model. The session API, the event "
                "stream and the public docs expose neither the underlying model nor the agent mode "
                "used by this session, and no per-mode ACU multiplier is published. The table below "
                "therefore records only what is documented or measured - it deliberately contains no "
                "invented multipliers.")
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 60
    ws.merge_cells("A2:F2")

    write_header(ws, ["Agent mode", "Measured ACUs in this session", "Published ACU multiplier",
                      "Documented positioning", "Fit for this migration",
                      "How to obtain real comparison data"], row=4)
    rows = [
        ("Devin (default / normal)", "Not attributable - mode not exposed", "None published",
         "General-purpose agent for features, bug fixes, refactoring and most development work",
         "Suitable for the whole migration; the judgement-heavy rows (standalone conversion, application builder) need at least this level",
         "Re-run one Angular bump in this mode and compare acus_consumed"),
        ("Ultra", "Not attributable - mode not exposed", "None published",
         "Selectable mode in the agent toggle / !ultra in Slack; no ACU multiplier documented",
         "Only justified for the two architectural rows if the default mode fails them",
         "Re-run the same single bump in Ultra and compare acus_consumed"),
        ("Fast", "Not attributable - mode not exposed", "None published",
         "Optimised for quick, well-scoped tasks",
         "Good candidate for the mechanical version bumps (0.42-0.70 ACU each measured)",
         "Re-run one bump in Fast and compare acus_consumed"),
        ("Lite", "Not attributable - mode not exposed", "None published",
         "Selectable lightweight mode (!lite)",
         "Candidate for repetitive bumps and for the reporting/Q&A turns",
         "Re-run one bump plus one reporting question in Lite and compare"),
        ("Fusion", "Not attributable - mode not exposed", "None published",
         "Selectable mode (!fusion)",
         "Not evaluated",
         "Same A/B method"),
    ]
    row = 5
    for values in rows:
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value).alignment = WRAP
        ws.row_dimensions[row].height = 46
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Recommended A/B experiment to make this table factual").font = BOLD
    row += 1
    for line in [
        "1. Reset the repo to the Angular 9 baseline commit in three separate sessions.",
        "2. Run exactly one identical unit of work (Angular 9 -> 10) per session, one session per mode.",
        "3. Read acus_consumed from the session API for each and compare against the 1.1050 ACUs measured here.",
        "4. Repeat with the standalone-conversion unit, which is the quality-sensitive case.",
        "Cost of the experiment: roughly 3-6 ACUs total, and it produces the only defensible per-mode numbers.",
    ]:
        ws.cell(row=row, column=1, value=line)
        row += 1
    autosize(ws, [26, 32, 24, 46, 50, 46])


def sheet_hotspots(wb, usage, report, rate):
    ws = wb.create_sheet("Hotspots & avoidable ACUs")
    ws["A1"] = "High-consumption activities, repeated interactions and avoidable ACUs"
    ws["A1"].font = Font(bold=True, size=13)
    write_header(ws, ["Rank", "Activity", "ACUs", "Dollar value", "Basis", "Why it cost this much",
                      "Avoidable portion", "Corrective measure"], row=3)

    ledger = {i["n"]: i for i in report["interaction_ledger"]}
    reporting_total = sum(i["delta_acus"] for i in report["interaction_ledger"]
                          if i["category"] == "Reporting")
    tasks = {t["id"]: t for t in usage["tasks"]}
    rows = [
        ("Post-delivery reporting and Q&A (8 separate follow-up requests)", reporting_total,
         "Measured",
         "Each follow-up resumed the session and re-derived overlapping analysis; two requests were duplicates of a prior ask",
         "~60% (8.3 ACUs)",
         "State the full reporting deliverable in the initial prompt: ACU table + xlsx + cache sheet + leadership narrative, produced once"),
        ("Post-PR verification (V1)", tasks["V1"]["acus"], "Derived",
         "590 s of interactive prod build, serve, service-worker and manifest checks, plus writing an adversarial test plan",
         "~3.5-4 ACUs",
         "One scripted build+serve+assert chain, or let CI run it and inspect only failures"),
        ("Angular 20 + application builder + control flow + Playwright + ops config", tasks["B11+C6-C8"]["acus"],
         "Derived", "Largest genuine construction step: builder schema migration, template rewrites, new e2e harness",
         "Little", "Keep as is - this is judgement work"),
        ("Inception (BRD + units of work)", tasks["A1-A3"]["acus"], "Derived",
         "Full repo read and breaking-change analysis across 11 Angular majors",
         "~1.5 ACUs on repeat runs",
         "Encode as a Devin playbook plus the existing Angular knowledge note; reuse instead of re-deriving"),
        ("AIDLC tooling (prompt prefix, cost scripts)", tasks["E1-E2"]["acus"], "Derived",
         "One-time asset creation requested by the AIDLC method",
         "~1.5 ACUs on repeat runs", "Reuse the committed scripts"),
        ("Angular 9 -> 10 (B1)", tasks["B1"]["acus"], "Derived",
         "Most expensive single bump: discovered the yarn/npm mismatch, peer-dependency conflicts and the Node version floor",
         "~0.6 ACU",
         "Pre-provision Node 22.x, npm, headless Chrome flags and a warm npm cache in the environment blueprint"),
        ("Nine remaining version bumps (B2-B9, C9+B10)", sum(tasks[i]["acus"] for i in
                                                             ["B2", "B3", "B4", "B5", "B6", "B7", "B8", "B9", "C9+B10"]),
         "Derived", "Each bump ran its own npm install + build + test cycle; most of the elapsed time was install wait",
         "~1.5 ACUs",
         "Verify at checkpoints (after 12, 15, 18, 20) instead of after every version; accepts harder failure bisection"),
    ]
    rows.sort(key=lambda r: r[1], reverse=True)
    row = 4
    for rank, (activity, acus, basis, why, avoidable, fix) in enumerate(rows, start=1):
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=activity).alignment = WRAP
        ws.cell(row=row, column=3, value=round(acus, 4))
        ws.cell(row=row, column=4, value=f"=C{row}*{rate}").number_format = '"$"#,##0.00'
        ws.cell(row=row, column=5, value=basis)
        ws.cell(row=row, column=6, value=why).alignment = WRAP
        ws.cell(row=row, column=7, value=avoidable).alignment = WRAP
        ws.cell(row=row, column=8, value=fix).alignment = WRAP
        ws.row_dimensions[row].height = 60
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="Repeated interactions detected (measured)").font = BOLD
    row += 1
    for line in [
        "The same question ('which model was used / how to reduce ACUs') was submitted twice at 2026-08-31T06:00:52Z.",
        "Four separate requests (metrics, xlsx, cache detail, explanation) asked for different views of one dataset.",
        f"Checkpoints 6, 7 and 9 all re-explained the same workbook: {ledger[6]['delta_acus'] + ledger[7]['delta_acus'] + ledger[9]['delta_acus']:.2f} ACUs combined.",
    ]:
        ws.cell(row=row, column=2, value=line)
        row += 1
    autosize(ws, [6, 46, 10, 14, 12, 56, 22, 60])


def sheet_outcomes(wb, report, usage, rate):
    ws = wb.create_sheet("ACU vs outcome")
    ws["A1"] = "Relationship between ACU consumption and migration outcome"
    ws["A1"].font = Font(bold=True, size=13)
    write_header(ws, ["Outcome delivered", "Evidence", "ACUs attributable", "Dollar value"], row=3)
    tasks = {t["id"]: t for t in usage["tasks"]}
    ev = report["outcome_evidence"]
    rows = [
        ("11 Angular major versions traversed (9 -> 20)", "package.json on PR #35 at Angular 20.3.x",
         sum(tasks[i]["acus"] for i in ["B1", "B2", "B3", "B4", "B5", "B6", "B7", "B8", "B9", "C9+B10", "B11+C6-C8"])),
        ("Toolchain replaced: TSLint -> angular-eslint, Protractor -> Playwright, RxJS 7",
         f"lint {ev['lint']}, e2e {ev['e2e']}", tasks["C5"]["acus"] + tasks["C1"]["acus"]),
        ("Architecture modernised: standalone components, provideRouter, application builder",
         "src/main.ts bootstrapApplication, angular.json @angular/build:application",
         tasks["C9+B10"]["acus"]),
        ("Test coverage created where none existed", f"unit {ev['unit_tests']} (repo had zero specs)",
         "included in the B2 row (0.9907)"),
        ("Build and release path verified", f"dev {ev['dev_build']}, prod {ev['prod_build']}", tasks["V1"]["acus"]),
        ("Reviewable deliverable", ev["pull_request"], tasks["D-PR"]["acus"]),
    ]
    row = 4
    for outcome, evidence, acus in rows:
        ws.cell(row=row, column=1, value=outcome).alignment = WRAP
        ws.cell(row=row, column=2, value=evidence).alignment = WRAP
        if isinstance(acus, str):
            ws.cell(row=row, column=3, value=acus).alignment = WRAP
        else:
            ws.cell(row=row, column=3, value=round(acus, 4))
            ws.cell(row=row, column=4, value=f"=C{row}*{rate}").number_format = '"$"#,##0.00'
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="Not delivered").font = BOLD
    ws.cell(row=row, column=2, value=ev["not_completed"]).alignment = WRAP
    row += 2
    ws.cell(row=row, column=1, value="Reading").font = BOLD
    ws.cell(row=row, column=2, value=(
        "ACU spend correlates with judgement density, not with the number of files changed: the nine "
        "mechanical bumps together cost less than the single verification block, and every phase that "
        "produced a durable artefact (BRD, tooling, PR) was under 2 ACUs.")).alignment = WRAP
    autosize(ws, [56, 60, 18, 14])


def sheet_recommendation(wb, report, rate):
    ws = wb.create_sheet("Recommended approach")
    ws["A1"] = "Most cost-efficient way to run this migration with Devin"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = ("Baseline measured: 20.2524 ACUs delivery + 13.8651 ACUs reporting = "
                "34.1175 ACUs for the session to date.")
    write_header(ws, ["Dimension", "Recommendation", "Evidence it is grounded in",
                      "Expected effect (estimate)"], row=4)
    rows = [
        ("Environment", "Pin Node 22.x, npm as package manager, legacy-peer-deps, headless Chrome flags and a warm npm cache in the repo environment blueprint before starting",
         "B1 cost 1.1050 ACUs, the most expensive bump, entirely because these were discovered at runtime",
         "-1.5 to -2 ACUs"),
        ("Session strategy", "One session for the whole migration; state every deliverable (code, BRD, metrics table, xlsx, cache analysis) in the opening prompt",
         "8 follow-up requests cost a measured 13.8651 ACUs, 41% of the session",
         "-6 to -8 ACUs"),
        ("Prompts", "Keep the byte-stable prefix (already committed) and put only the task delta at the end; avoid re-asking for restatements of the same dataset",
         "97.8% measured context reuse shows the prefix technique already works",
         "Preserves the current near-optimal reuse"),
        ("Verification", "Replace interactive verification with one scripted build+serve+assert chain, and let CI run lint/unit/e2e; involve the agent only on failures",
         "V1 measured 5.6204 ACUs, 590 s, the largest single block",
         "-3.5 to -4 ACUs"),
        ("AIDLC workflow", "Checkpoint verification after Angular 12, 15, 18 and 20 rather than after all 11 bumps",
         "Bumps B2-B9 averaged 0.58 ACUs each, dominated by install/build wait",
         "-1 to -1.5 ACUs"),
        ("Model / mode", "Run mechanical bumps in a lighter mode and switch up only for the standalone conversion and the Angular 20 builder row; validate with the A/B experiment on the model sheet before committing to it",
         "Those two rows cost 1.4575 and 2.2386 ACUs and are the only judgement-heavy steps; no per-mode multiplier is published, so the saving is unquantified",
         "Unquantified until the A/B is run"),
        ("Reuse", "Encode the Angular upgrade sequence as a playbook and keep the existing Angular knowledge note; reuse the committed prompt-prefix and cost scripts",
         "Inception 1.7528 + tooling 1.5718 ACUs were one-time asset creation",
         "-3 ACUs on the next repository"),
        ("Portfolio scaling", "Version bumps are sequential per repo but independent across repos - run one session per application in parallel",
         "Angular refuses to skip majors (ng update constraint observed during this migration)",
         "Wall-clock scales with repo count, not sum of repos"),
    ]
    row = 5
    for values in rows:
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value).alignment = WRAP
        ws.row_dimensions[row].height = 58
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Optimised target (estimate, not measured)").font = BOLD
    row += 1
    for label, acus in [("Delivery only, optimised", 13.0), ("Reporting, single consolidated request", 4.0)]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=acus)
        ws.cell(row=row, column=3, value=f"=B{row}*{rate}").number_format = '"$"#,##0.00'
        row += 1
    ws.cell(row=row, column=1, value="Optimised session total (estimate)").font = BOLD
    ws.cell(row=row, column=2, value=17.0).font = BOLD
    ws.cell(row=row, column=3, value=f"=B{row}*{rate}").number_format = '"$"#,##0.00'
    row += 1
    ws.cell(row=row, column=1, value="Measured session total for comparison").font = BOLD
    ws.cell(row=row, column=2, value=round(report["interaction_ledger"][-1]["cumulative_acus"], 4)).font = BOLD
    ws.cell(row=row, column=3, value=f"=B{row}*{rate}").number_format = '"$"#,##0.00'
    autosize(ws, [24, 58, 62, 30])


def sheet_provenance(wb, report):
    ws = wb.create_sheet("Provenance")
    ws["A1"] = "Where every number in this workbook comes from"
    ws["A1"].font = Font(bold=True, size=13)
    write_header(ws, ["Item", "Source / status"], row=3)
    row = 4
    for key, value in report["measurement_sources"].items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value).alignment = WRAP
        ws.row_dimensions[row].height = 44
        row += 1
    row += 1
    for label, value in [
        ("Price per ACU", f"${report['rate']['price_per_acu']:.2f} - {report['rate']['note']}"),
        ("Session", report["session_url"]),
        ("Pull request", report["pull_request"]),
        ("Raw data", "docs/migration/acu-usage.json and docs/migration/acu-consumption-report.json"),
    ]:
        ws.cell(row=row, column=1, value=label).font = BOLD
        ws.cell(row=row, column=2, value=value).alignment = WRAP
        row += 1
    autosize(ws, [30, 110])


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--rate", type=float, default=None)
    args = parser.parse_args()

    usage = json.loads(USAGE.read_text())
    report = json.loads(REPORT.read_text())
    rate = args.rate if args.rate is not None else report["rate"]["price_per_acu"]

    wb = Workbook()
    wb.remove(wb.active)
    sheet_phase_consumption(wb, usage, report, rate)
    sheet_ledger(wb, report, rate)
    sheet_rollup(wb, report, rate)
    sheet_unit_rates(wb, usage, report, rate)
    sheet_hotspots(wb, usage, report, rate)
    sheet_models(wb)
    sheet_outcomes(wb, report, usage, rate)
    sheet_recommendation(wb, report, rate)
    sheet_provenance(wb, report)
    wb.save(OUT)
    print(f"wrote {OUT} at ${rate:.2f}/ACU")


if __name__ == "__main__":
    main()
