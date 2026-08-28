#!/usr/bin/env python3
"""Render docs/migration/acu-usage.json as an AIDLC-style xlsx cost sheet.

Usage: python3 scripts/aidlc-acu-xlsx.py [--rate 2.25] [--out docs/migration/aidlc-acu-metrics.xlsx]
"""
import argparse
import json
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--rate", type=float, default=None, help="USD per ACU")
    ap.add_argument("--out", default="docs/migration/aidlc-acu-metrics.xlsx")
    args = ap.parse_args()

    usage = json.loads((ROOT / "docs/migration/acu-usage.json").read_text())
    rate = args.rate if args.rate is not None else usage["rate"]["price_per_acu"]

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style="thin", color="B4C6E7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.active
    ws.title = "ACU Cost"
    ws["A1"] = "Angular 9 → 20 migration — Devin ACU cost (AIDLC)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Rate: ${rate:.2f} per ACU"
    ws["A3"] = f"Session: {usage['session_url']}"
    ws["A4"] = f"Pull request: {usage['pull_request']}"

    header = ["Unit", "AIDLC task", "Start (UTC)", "End (UTC)", "Duration (s)", "ACUs", "Cost (USD)"]
    ws.append([])
    ws.append(header)
    hrow = ws.max_row
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=hrow, column=c)
        cell.fill, cell.font, cell.border = head_fill, head_font, border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for t in usage["tasks"]:
        ws.append([t["id"], t["task"], t["start_utc"], t["end_utc"], t["duration_seconds"], t["acus"], round(t["acus"] * rate, 2)])

    first, last = hrow + 1, ws.max_row
    ws.append(["Total", "Angular 9 → 20 migration", "", "", f"=SUM(E{first}:E{last})", f"=SUM(F{first}:F{last})", f"=SUM(G{first}:G{last})"])
    for c in range(1, len(header) + 1):
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=c).fill = total_fill

    for row in ws.iter_rows(min_row=first, max_row=ws.max_row, max_col=len(header)):
        for cell in row:
            cell.border = border
            if cell.column_letter == "F":
                cell.number_format = "0.0000"
            if cell.column_letter == "G":
                cell.number_format = '"$"#,##0.00'
            if cell.column_letter == "B":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col, width in zip("ABCDEFG", (12, 62, 12, 12, 13, 11, 13)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = f"A{first}"

    sm = wb.create_sheet("Summary")
    measured = usage["measured"]
    rows = [
        ("Metric", "Value"),
        ("Repository", usage["repository"]),
        ("Session", usage["session_url"]),
        ("Pull request", usage["pull_request"]),
        ("Metric type", measured["metric"]),
        ("Source", measured["source"]),
        ("ACUs at session start", measured["acus_at_session_start"]),
        ("ACUs at end of migration work", measured["acus_at_end_of_migration_work"]),
        ("Migration ACUs (measured)", measured["migration_acus"]),
        ("Migration window (UTC)", " → ".join(measured["migration_window_utc"])),
        ("Billable wall-clock (minutes)", measured["billable_wall_clock_minutes"]),
        ("Price per ACU (USD)", rate),
        ("Total cost (USD)", round(measured["migration_acus"] * rate, 2)),
        ("Attribution method", usage["attribution"]["method"]),
        ("Confidence", usage["attribution"]["confidence"]),
    ]
    for r in rows:
        sm.append(list(r))
    for c in range(1, 3):
        cell = sm.cell(row=1, column=c)
        cell.fill, cell.font = head_fill, head_font
    sm.column_dimensions["A"].width = 32
    sm.column_dimensions["B"].width = 110
    for row in sm.iter_rows(min_row=2, max_row=sm.max_row, min_col=2, max_col=2):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
    sm.cell(row=13, column=2).number_format = '"$"#,##0.00'

    cache = usage.get("prompt_cache")
    if cache:
        cs = wb.create_sheet("Prompt cache")
        cs.append(["Prompt cache / context reuse"])
        cs["A1"].font = Font(bold=True, size=14)
        cs.append([cache["note"]])
        cs.append([f"Source: {cache['source']}"])
        cs.append([])
        cs.append(["Scope", "Iterations", "Prompt tokens processed", "Cache-read tokens", "Cache-creation tokens", "Cache hit rate"])
        for c in range(1, 7):
            cell = cs.cell(row=cs.max_row, column=c)
            cell.fill, cell.font, cell.border = head_fill, head_font, border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for label, key in (("Measured window", "measured_window"), ("Full session (extrapolated)", "extrapolated_full_session")):
            d = cache[key]
            cs.append([label, f"{d['iterations'][0]}-{d['iterations'][1]}", d["prompt_tokens"], d["cache_read_tokens"], d["cache_creation_tokens"], d["hit_rate_pct"] / 100])
            for c in range(1, 7):
                cs.cell(row=cs.max_row, column=c).border = border
            cs.cell(row=cs.max_row, column=6).number_format = "0.0%"
            for c in (3, 4, 5):
                cs.cell(row=cs.max_row, column=c).number_format = "#,##0"
        cs.append([])
        cs.append(["Agent iteration", "Context tokens", "Cache-read tokens", "Cache-creation tokens"])
        for c in range(1, 5):
            cell = cs.cell(row=cs.max_row, column=c)
            cell.fill, cell.font, cell.border = head_fill, head_font, border
        prev = 0
        for it, tokens in sorted(cache["context_tokens_per_iteration"].items(), key=lambda kv: int(kv[0])):
            cs.append([int(it), tokens, prev, tokens - prev])
            for c in range(1, 5):
                cs.cell(row=cs.max_row, column=c).border = border
                cs.cell(row=cs.max_row, column=c).number_format = "#,##0"
            prev = tokens
        for col, width in zip("ABCDEF", (28, 18, 24, 22, 24, 16)):
            cs.column_dimensions[col].width = width

    out = ROOT / args.out
    wb.save(out)
    print(f"wrote {out} ({out.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
