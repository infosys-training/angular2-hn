#!/usr/bin/env python3
"""DEVIN AIDLC token & ACU workbook - phase-merged version.

Layout follows the reviewed workbook:
  AIDLC Phases | Task | Model | Token | Input | Cache Write | Cache Read |
  Output | ACUs | Basis | Dollar Value | What is covered

Phase rows are merged: Operation (PR) and Validation are reported inside
Construction, and Governance / AIDLC tooling is marked optional and placed last.

Sources (all captured in this session, all in this repo):
  docs/migration/acu-usage.json               measured ACUs + per-task wall clock
  docs/migration/acu-consumption-report.json  measured per-interaction ACU deltas
  docs/migration/token-telemetry.json         measured context_growth_update samples

Usage: python3 scripts/devin-aidlc-acu-metrics-v2.py [--rate 2.25]
"""
import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
USAGE = ROOT / "docs/migration/acu-usage.json"
REPORT = ROOT / "docs/migration/acu-consumption-report.json"
TELEMETRY = ROOT / "docs/migration/token-telemetry.json"
OUT = ROOT / "docs/migration/DEVIN-AIDLC-ACU-metrics-v2.xlsx"

HEAD = PatternFill("solid", fgColor="9BC2E6")
SUB = PatternFill("solid", fgColor="DDEBF7")
WARN = PatternFill("solid", fgColor="FFF2CC")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
MODEL_CELL = "Devin agent (mode not exposed)"

CLAUDE_RATES = {
    "Opus": {"input": 5.0, "cache_write": 6.25, "cache_read": 0.5, "output": 25.0},
    "Sonnet 4.6": {"input": 3.0, "cache_write": 2.5, "cache_read": 0.3, "output": 15.0},
}

# Which task ids roll up into which reported phase row.
CONSTRUCTION_IDS = {"B1", "B2", "B3", "B4", "B5", "B6", "B7", "B8", "B9",
                    "C1", "C5", "C9+B10", "B11+C6-C8", "D-PR", "V1"}
INCEPTION_IDS = {"A1-A3"}
GOVERNANCE_IDS = {"E1-E2"}

COVERED = {
    "Inception": (
        "Understood the existing app and planned the migration: BRD of current vs target versions and every "
        "breaking change, broken into one unit of work per version with acceptance criteria."),
    "Construction": (
        "Did the migration: 11 major Angular upgrades one at a time, modernised the code and toolchain "
        "(RxJS 7, ESLint, standalone, new control flow, new builder, Playwright tests) and updated the CI, "
        "hosting and docs config. Includes Operation (raised PR #35) and Validation (builds, lint, unit and "
        "e2e tests green after each step)."),
    "Governance / AIDLC tooling (optional)": (
        "Optional. Reusable AIDLC instrumentation - the stable prompt prefix and the cost/usage reporting "
        "scripts. Not needed to migrate the app, and reusable on future migrations."),
}

BASIS = {
    "Inception": "ACUs derived (wall-clock split of the measured 20.2524 total); tokens = ACUs x measured construction token rate",
    "Construction": "ACUs derived (wall-clock split of the measured 20.2524 total); tokens = ACUs x measured construction token rate",
    "Governance / AIDLC tooling (optional)": "ACUs derived (wall-clock split of the measured 20.2524 total); tokens = ACUs x measured construction token rate",
}


def write_header(ws, headers, row=1, fill=HEAD):
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=text)
        c.font = BOLD
        c.fill = fill
        c.alignment = WRAP
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def token_rates(region, acus):
    return {
        "cache_read": region["cache_read_tokens"] / acus,
        "cache_write": region["cache_creation_tokens"] / acus,
        "input": region["input_tokens"] / acus,
        "output": region["output_tokens"] / acus,
    }


def phase_acus(usage):
    by_id = {t["id"]: t["acus"] for t in usage["tasks"]}
    return [
        ("Inception", sum(v for k, v in by_id.items() if k in INCEPTION_IDS)),
        ("Construction", sum(v for k, v in by_id.items() if k in CONSTRUCTION_IDS)),
        ("Governance / AIDLC tooling (optional)",
         sum(v for k, v in by_id.items() if k in GOVERNANCE_IDS)),
    ]


TASK_CELL = {
    "Inception": "A1-A3  BRD, units of work, acceptance criteria",
    "Construction": "B1-B11, C1-C9, D-PR, V1  Angular 9 -> 20 upgrade, cross-cutting refactors, PR and verification",
    "Governance / AIDLC tooling (optional)": "E1-E2  Prompt-prefix generator and AIDLC cost tooling",
}


def sheet_main(wb, usage, rate, rates):
    ws = wb.create_sheet("AIDLC token & ACU")
    ws["A1"] = "DEVIN - Angular 9 -> 20 migration: AIDLC token & ACU consumption"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (f"Devin bills ACUs, not tokens. Dollar Value = ACUs x ${rate:.2f}/ACU. "
                "Token / Input / Cache Write / Cache Read / Output are measured context telemetry shown for "
                "parity with the Claude sheet - they carry no price. Operation (PR) and Validation are reported "
                "inside Construction; Governance is optional tooling and is listed last.")
    ws["A2"].alignment = WRAP
    ws["A3"] = f'Session devin-b089339539c84cfb8fe4913619e49d66 | {usage["repository"]} | PR {usage["pull_request"]}'

    headers = ["AIDLC Phases", "Task", "Model", "Token", "Input", "Cache Write", "Cache Read",
               "Output", "ACUs", "Basis", "Dollar Value", "What is covered"]
    write_header(ws, headers, row=5)

    r = rates["construction"]
    row = 6
    first = row
    for phase, acus in phase_acus(usage):
        ws.cell(row=row, column=1, value=phase).alignment = WRAP
        ws.cell(row=row, column=2, value=TASK_CELL[phase]).alignment = WRAP
        ws.cell(row=row, column=3, value=MODEL_CELL).alignment = WRAP
        ws.cell(row=row, column=4, value=f"=G{row}+F{row}").number_format = "#,##0"
        ws.cell(row=row, column=5, value=round(acus * r["input"])).number_format = "#,##0"
        ws.cell(row=row, column=6, value=round(acus * r["cache_write"])).number_format = "#,##0"
        ws.cell(row=row, column=7, value=round(acus * r["cache_read"])).number_format = "#,##0"
        ws.cell(row=row, column=8, value=round(acus * r["output"])).number_format = "#,##0"
        ws.cell(row=row, column=9, value=round(acus, 4))
        ws.cell(row=row, column=10, value=BASIS[phase]).alignment = WRAP
        ws.cell(row=row, column=11, value=f"=I{row}*{rate}").number_format = '"$"#,##0.00'
        ws.cell(row=row, column=12, value=COVERED[phase]).alignment = WRAP
        row += 1
    last = row - 1

    ws.cell(row=row, column=1, value="MIGRATION TOTAL").font = BOLD
    for col in (4, 5, 6, 7, 8, 9):
        L = get_column_letter(col)
        c = ws.cell(row=row, column=col, value=f"=SUM({L}{first}:{L}{last})")
        c.font = BOLD
        c.number_format = "0.0000" if col == 9 else "#,##0"
    c = ws.cell(row=row, column=11, value=f"=I{row}*{rate}")
    c.font = BOLD
    c.number_format = '"$"#,##0.00'
    ws.cell(row=row, column=10, value="Migration ACU total is MEASURED: 20.2524 ACUs").alignment = WRAP
    for col in range(1, 13):
        ws.cell(row=row, column=col).fill = SUB
    total_row = row

    row += 1
    ws.cell(row=row, column=1, value="Delivery total excluding optional governance tooling").alignment = WRAP
    ws.cell(row=row, column=9, value=f"=I{total_row}-I{last}")
    ws.cell(row=row, column=11, value=f"=I{row}*{rate}").number_format = '"$"#,##0.00'
    ws.cell(row=row, column=12, value="Cost of the migration alone, with the reusable tooling "
                                      "excluded.").alignment = WRAP

    row += 2
    ws.cell(row=row, column=1, value="Note").font = BOLD
    ws.cell(row=row, column=2, value=(
        "Post-delivery reporting and Q&A (13.8651 measured ACUs / $31.20) is deliberately NOT in this table: it "
        "is not migration work. It is shown separately in 'If tokens were billed' and in the review script.")).alignment = WRAP
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)

    autosize(ws, [24, 46, 24, 13, 11, 13, 14, 11, 10, 40, 13, 84])
    ws.row_dimensions[5].height = 30
    for rr in range(first, last + 1):
        ws.row_dimensions[rr].height = 62
    return ws


def sheet_phases(wb, usage, rate):
    ws = wb.create_sheet("AIDLC Phases")
    ws["A1"] = "ACU consumption per AIDLC phase (detail behind the merged rows)"
    ws["A1"].font = Font(bold=True, size=12)
    write_header(ws, ["AIDLC phase", "Reported inside", "ACUs", "Dollar value", "% of migration",
                      "Basis", "What it covers"], row=3)
    by_id = {t["id"]: (t["acus"], t["task"]) for t in usage["tasks"]}
    detail = [
        ("Inception", "own row", sum(by_id[i][0] for i in INCEPTION_IDS),
         "Repo read, BRD, units of work, acceptance criteria"),
        ("Construction (upgrades + refactors)", "Construction",
         sum(v for k, (v, _) in by_id.items() if k in CONSTRUCTION_IDS - {"D-PR", "V1"}),
         "11 Angular major bumps, RxJS 7, angular-eslint, standalone, control flow, application builder, "
         "Playwright, ops config"),
        ("Operation", "Construction", by_id["D-PR"][0], "PR #35 raised with the migration summary"),
        ("Validation", "Construction", by_id["V1"][0],
         "Production build, service-worker/manifest serve check, adversarial test plan, browser-testing handoff"),
        ("Governance / AIDLC tooling (optional)", "own row, last",
         sum(by_id[i][0] for i in GOVERNANCE_IDS),
         "Prompt-prefix generator and cost tooling - reusable asset, not required by the migration"),
    ]
    r = 4
    first = r
    for phase, inside, acus, covers in detail:
        ws.cell(row=r, column=1, value=phase).alignment = WRAP
        ws.cell(row=r, column=2, value=inside)
        ws.cell(row=r, column=3, value=round(acus, 4))
        ws.cell(row=r, column=4, value=f"=C{r}*{rate}").number_format = '"$"#,##0.00'
        ws.cell(row=r, column=6, value="Derived (wall-clock split of the measured total)").alignment = WRAP
        ws.cell(row=r, column=7, value=covers).alignment = WRAP
        r += 1
    last = r - 1
    ws.cell(row=r, column=1, value="MIGRATION TOTAL").font = BOLD
    ws.cell(row=r, column=3, value=f"=SUM(C{first}:C{last})").font = BOLD
    ws.cell(row=r, column=4, value=f"=C{r}*{rate}").number_format = '"$"#,##0.00'
    ws.cell(row=r, column=6, value="MEASURED: 20.2524 ACUs").font = BOLD
    total_row = r
    for rr in range(first, last + 1):
        ws.cell(row=rr, column=5, value=f"=C{rr}/$C${total_row}").number_format = "0.0%"
    ws.cell(row=r, column=5, value=f"=C{r}/$C${total_row}").number_format = "0.0%"
    autosize(ws, [36, 18, 10, 13, 14, 34, 70])
    return ws


def sheet_rate_card(wb, tel, rate, rates):
    ws = wb.create_sheet("Rate card & token rates")
    ws["A1"] = "What Devin charges for"
    ws["A1"].font = Font(bold=True, size=12)
    write_header(ws, ["Billable unit", "Price", "Billed?", "Note"], row=3)
    rows = [
        ("ACU (Agent Compute Unit)", f"${rate:.2f} per ACU", "YES",
         "Only billable unit. Enterprise price is set on the order form; $2.25 is the list-price assumption."),
        ("Input tokens", "$0.00", "No", "Not billed and not exposed as a billing field by Devin."),
        ("Cache write tokens", "$0.00", "No", "Not billed. Prompt caching shows up only as fewer ACUs."),
        ("Cache read tokens", "$0.00", "No", "Not billed."),
        ("Output tokens", "$0.00", "No", "Not billed."),
    ]
    r = 4
    for a, b, c, d in rows:
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        ws.cell(row=r, column=3, value=c).font = BOLD
        ws.cell(row=r, column=4, value=d).alignment = WRAP
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Measured Devin token rates (tokens per ACU)").font = Font(bold=True, size=12)
    r += 1
    write_header(ws, ["Region", "Telemetry window", "Iterations", "ACUs in window", "Cache read / ACU",
                      "Cache write / ACU", "Input / ACU", "Output / ACU"], row=r)
    r += 1
    for name, key, acus in (("Construction (migration)", "construction_measured", rates["construction_acus"]),
                            ("Reporting (post-delivery Q&A)", "reporting_measured", rates["reporting_acus"])):
        reg = tel["regions"][key]
        k = "construction" if key.startswith("construction") else "reporting"
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=f'{reg["window_utc"][0]} -> {reg["window_utc"][1]} '
                                       f'(iterations {reg["iterations"][0]}-{reg["iterations"][1]}, '
                                       f'{reg["samples"]} samples)').alignment = WRAP
        ws.cell(row=r, column=3, value=reg["iteration_steps"])
        ws.cell(row=r, column=4, value=round(acus, 4))
        for col, kk in ((5, "cache_read"), (6, "cache_write"), (7, "input"), (8, "output")):
            ws.cell(row=r, column=col, value=round(rates[k][kk])).number_format = "#,##0"
        r += 1
    r += 1
    ws.cell(row=r, column=1, value=(
        "Every token cell on the 'AIDLC token & ACU' sheet = that phase's ACUs x the construction rate above. "
        "The rates are measured, not assumed: context telemetry of a fully sampled window divided by the ACUs "
        "consumed in the same window.")).alignment = WRAP
    autosize(ws, [30, 52, 12, 15, 18, 18, 14, 14])


def sheet_provenance(wb, usage, report, tel, rate, rates):
    ws = wb.create_sheet("Provenance & formulas")
    write_header(ws, ["Figure", "Value", "How it was obtained", "Basis"], row=1)
    mig = usage["measured"]["migration_acus"]
    rep = sum(e["delta_acus"] for e in report["interaction_ledger"] if e["category"] == "Reporting")
    A = tel["regions"]["construction_measured"]
    rows = [
        ("Migration ACUs", round(mig, 4),
         "Devin session API: acus_consumed at end of migration work minus 0 at session start "
         "(2026-08-28 17:06:16Z -> 17:52:44Z)", "MEASURED"),
        ("Post-delivery reporting ACUs", round(rep, 4),
         "Sum of acu_consumption_at_last_user_interaction deltas for the 7 reporting requests "
         "(excluded from the phase table)", "MEASURED"),
        ("Price per ACU", rate, "Devin list-price assumption; enterprise rate is set on your order form. "
         "Re-run with --rate to change it", "ASSUMPTION"),
        ("Dollar Value column", "ACUs x rate", "Live formula (=I<row>*rate); totals are =SUM() over the phase rows",
         "MEASURED x ASSUMPTION"),
        ("Phase ACUs", "3 merged rows", "Measured migration total split pro rata by wall-clock duration of the "
         "17 underlying tasks, then summed per phase. Devin exposes no per-task ACU counter", "DERIVED"),
        ("Cache read tokens", f'{A["cache_read_tokens"]:,} in the construction window',
         "context_growth_update: context resident before each iteration, i.e. the prefix a cache serves",
         "MEASURED (proxy for cache_read_input_tokens)"),
        ("Cache write tokens", f'{A["cache_creation_tokens"]:,} in the construction window',
         "Positive delta of current_context_tokens between iterations", "MEASURED (proxy)"),
        ("Input tokens", f'{A["input_tokens"]:,} in the construction window',
         "Share of new tokens attributable to tool results (delta total_tool_output_bytes / bytes-per-token)",
         "MEASURED (proxy)"),
        ("Output tokens", f'{A["output_tokens"]:,} in the construction window',
         "New tokens minus tool-result tokens = model-generated text and tool calls", "MEASURED (proxy)"),
        ("Token rate per ACU", "see 'Rate card & token rates'",
         f'Construction: window tokens / {rates["construction_acus"]:.4f} ACUs consumed in the same window',
         "MEASURED"),
        ("Model / agent mode", "Not exposed",
         "The Devin session API returns no model identity or agent mode for a completed session, and Devin "
         "publishes no per-mode ACU multiplier. Measure it with a controlled A/B re-run instead", "NOT AVAILABLE"),
        ("Migration outcome", "PR #35, lint clean, 6/6 unit specs, dev+prod build, 3/3 Playwright e2e",
         "Local command output on branch devin/1787936828-ng9-to-ng20-migration", "MEASURED"),
        ("Browser end-to-end run", "Not completed",
         "The persistent testing-agent run was suspended with usage_limit_exceeded; no recording exists",
         "NOT COMPLETED"),
    ]
    r = 2
    for a, b, c, d in rows:
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b).alignment = WRAP
        ws.cell(row=r, column=3, value=c).alignment = WRAP
        ws.cell(row=r, column=4, value=d).alignment = WRAP
        r += 1
    autosize(ws, [26, 34, 78, 34])


def sheet_cache(wb, tel):
    ws = wb.create_sheet("Cache efficiency")
    ws["A1"] = "Prompt cache / context reuse (measured)"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = tel["field_mapping"]["cache_read_tokens"]
    ws["A2"].alignment = WRAP
    write_header(ws, ["Window", "Iterations", "Prompt tokens processed", "Cache read (reused prefix)",
                      "Cache write (new tokens)", "Hit rate %", "New tokens / iteration", "Avg context size",
                      "Input tokens", "Output tokens"], row=4)
    r = 5
    for name, key in (("Construction, iterations 1-57", "construction_measured"),
                      ("Reporting, iterations 200-225", "reporting_measured")):
        reg = tel["regions"][key]
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=reg["iteration_steps"])
        for col, k in ((3, "prompt_tokens"), (4, "cache_read_tokens"), (5, "cache_creation_tokens")):
            ws.cell(row=r, column=col, value=reg[k]).number_format = "#,##0"
        ws.cell(row=r, column=6, value=reg["hit_rate_pct"])
        ws.cell(row=r, column=7, value=reg["new_tokens_per_iteration"]).number_format = "#,##0"
        ws.cell(row=r, column=8, value=reg["avg_context_tokens"]).number_format = "#,##0"
        ws.cell(row=r, column=9, value=reg["input_tokens"]).number_format = "#,##0"
        ws.cell(row=r, column=10, value=reg["output_tokens"]).number_format = "#,##0"
        r += 1
    r += 1
    for f in [
        "98.47% of tokens processed during construction were reused prefix - the byte-stable prompt prefix "
        "(package.json + angular.json + tsconfigs + file tree + conventions) did its job; the token side is saturated.",
        "Reporting turns write 3,234 new tokens per iteration vs 1,216 during construction: explaining the "
        "migration is 2.7x more token-dense per iteration than performing it.",
        f'Peak context {tel["peak_context_tokens"]:,} tokens; two compaction events (iterations 197->200 and '
        "221->225) discarded context, so negative deltas are excluded from cache-write.",
        "Devin never charges for any of this: cache efficiency reaches the invoice only as fewer ACUs.",
    ]:
        ws.cell(row=r, column=1, value=f).alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        r += 1
    autosize(ws, [46, 11, 15, 16, 16, 11, 14, 14, 13, 13])


def sheet_shadow(wb, usage, report, rate, rates):
    ws = wb.create_sheet("If tokens were billed")
    ws["A1"] = "Devin's token volume priced at the Claude rate card (comparison only - Devin bills none of this)"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = ("Rates below are the ones in the customer's Anthropic workbook. This sheet answers 'how does the "
                "ACU bill compare with a token bill for the same work?' It is a comparison, not an invoice.")
    ws["A2"].alignment = WRAP
    ws["A2"].fill = WARN
    write_header(ws, ["Model", "Input $/M", "Cache write $/M", "Cache read $/M", "Output $/M"], row=4)
    r = 5
    for name, rr in CLAUDE_RATES.items():
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=rr["input"])
        ws.cell(row=r, column=3, value=rr["cache_write"])
        ws.cell(row=r, column=4, value=rr["cache_read"])
        ws.cell(row=r, column=5, value=rr["output"])
        r += 1
    mig = usage["measured"]["migration_acus"]
    rep = sum(e["delta_acus"] for e in report["interaction_ledger"] if e["category"] == "Reporting")
    r += 1
    write_header(ws, ["Scope", "Input tokens", "Cache write tokens", "Cache read tokens", "Output tokens",
                      "Token cost @ Opus", "Token cost @ Sonnet 4.6", "Actual Devin bill (ACUs)"], row=r)
    r += 1
    scopes = [("Migration delivery", mig, "construction"),
              ("Post-delivery reporting", rep, "reporting"),
              ("Session total", None, None)]
    tot = {k: mig * rates["construction"][k] + rep * rates["reporting"][k]
           for k in ("input", "cache_write", "cache_read", "output")}
    for scope, acus, key in scopes:
        vals = tot if acus is None else {k: acus * rates[key][k]
                                         for k in ("input", "cache_write", "cache_read", "output")}
        acu_v = mig + rep if acus is None else acus
        ws.cell(row=r, column=1, value=scope)
        for col, k in ((2, "input"), (3, "cache_write"), (4, "cache_read"), (5, "output")):
            ws.cell(row=r, column=col, value=round(vals[k])).number_format = "#,##0"
        for col, model in ((6, "Opus"), (7, "Sonnet 4.6")):
            rr = CLAUDE_RATES[model]
            cost = (vals["input"] * rr["input"] + vals["cache_write"] * rr["cache_write"]
                    + vals["cache_read"] * rr["cache_read"] + vals["output"] * rr["output"]) / 1e6
            ws.cell(row=r, column=col, value=round(cost, 2)).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=8, value=round(acu_v * rate, 2)).number_format = '"$"#,##0.00'
        r += 1
    r += 1
    ws.cell(row=r, column=1, value=(
        "Not like-for-like: the token bill covers model inference only, the ACU bill covers inference plus the VM, "
        "the browser, every npm install and build, and the agent loop.")).alignment = WRAP
    ws.cell(row=r, column=1).fill = WARN
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    autosize(ws, [26, 16, 18, 18, 15, 17, 20, 22])


def sheet_samples(wb, tel):
    ws = wb.create_sheet("Raw telemetry")
    ws["A1"] = tel["source"]
    ws["A1"].alignment = WRAP
    write_header(ws, ["Timestamp (UTC)", "Iteration", "Context tokens", "Context bytes",
                      "Tool-output bytes (cum.)", "Tool-invocation bytes (cum.)",
                      "Delta context tokens vs previous sample"], row=3)
    r = 4
    prev = None
    for s in tel["samples"]:
        ws.cell(row=r, column=1, value=s["ts"][:19] + "Z")
        ws.cell(row=r, column=2, value=s["it"])
        ws.cell(row=r, column=3, value=s["ctx_tok"]).number_format = "#,##0"
        ws.cell(row=r, column=4, value=s["ctx_bytes"]).number_format = "#,##0"
        ws.cell(row=r, column=5, value=s["tool_out_bytes"]).number_format = "#,##0"
        ws.cell(row=r, column=6, value=s["tool_inv_bytes"]).number_format = "#,##0"
        if prev is not None:
            c = ws.cell(row=r, column=7, value=s["ctx_tok"] - prev)
            c.number_format = "#,##0"
            if s["ctx_tok"] < prev:
                c.fill = WARN
        prev = s["ctx_tok"]
        r += 1
    autosize(ws, [22, 11, 15, 14, 22, 24, 30])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--rate", type=float, default=2.25)
    args = ap.parse_args()

    usage = json.loads(USAGE.read_text())
    report = json.loads(REPORT.read_text())
    tel = json.loads(TELEMETRY.read_text())

    mig = usage["measured"]["migration_acus"]
    total_secs = sum(t["duration_seconds"] for t in usage["tasks"])
    construction_acus = mig * 741 / total_secs
    reporting_acus = report["interaction_ledger"][-1]["delta_acus"]
    rates = {
        "construction": token_rates(tel["regions"]["construction_measured"], construction_acus),
        "reporting": token_rates(tel["regions"]["reporting_measured"], reporting_acus),
        "construction_acus": construction_acus,
        "reporting_acus": reporting_acus,
    }

    wb = Workbook()
    wb.remove(wb.active)
    sheet_main(wb, usage, args.rate, rates)
    sheet_phases(wb, usage, args.rate)
    sheet_rate_card(wb, tel, args.rate, rates)
    sheet_provenance(wb, usage, report, tel, args.rate, rates)
    sheet_cache(wb, tel)
    sheet_shadow(wb, usage, report, args.rate, rates)
    sheet_samples(wb, tel)
    wb.save(OUT)
    print(f"wrote {OUT} at ${args.rate:.2f}/ACU")
    for phase, acus in phase_acus(usage):
        print(f"  {phase:44s} {acus:8.4f} ACUs  ${acus * args.rate:7.2f}")
    print(f"  {'TOTAL':44s} {sum(a for _, a in phase_acus(usage)):8.4f} ACUs")


if __name__ == "__main__":
    main()
