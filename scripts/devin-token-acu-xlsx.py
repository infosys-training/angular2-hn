#!/usr/bin/env python3
"""DEVIN token + ACU consumption workbook in the Claude AIDLC sheet layout.

Columns mirror the Claude workbook (AIDLC Phases | Task | Model | Token |
Dollar Value | Input | Cache Write | Cache Read | Output) and add the columns
Devin actually bills on (ACUs) plus the measurement basis of every row.

Sources (all read from this repo, all captured from this session):
  docs/migration/acu-usage.json               measured ACUs + per-task wall clock
  docs/migration/acu-consumption-report.json  measured per-interaction ACU deltas
  docs/migration/token-telemetry.json         measured context_growth_update samples

Usage:
  python3 scripts/devin-token-acu-xlsx.py [--rate 2.25]
"""
import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
USAGE = ROOT / "docs/migration/acu-usage.json"
REPORT = ROOT / "docs/migration/acu-consumption-report.json"
TELEMETRY = ROOT / "docs/migration/token-telemetry.json"
OUT = ROOT / "docs/migration/DEVIN-AIDLC-token-ACU-metrics.xlsx"

HEAD = PatternFill("solid", fgColor="9BC2E6")
SUB = PatternFill("solid", fgColor="DDEBF7")
WARN = PatternFill("solid", fgColor="FFF2CC")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

MODEL_CELL = "Devin agent (mode not exposed)"

# Claude workbook rate card (from the customer's Anthropic sheet), used only for
# the shadow comparison sheet - Devin does not bill tokens.
CLAUDE_RATES = {
    "Opus": {"input": 5.0, "cache_write": 6.25, "cache_read": 0.5, "output": 25.0},
    "Sonnet 4.6": {"input": 3.0, "cache_write": 2.5, "cache_read": 0.3, "output": 15.0},
}

PHASE_OF = {
    "A1-A3": "Inception",
    "C1": "Construction",
    "C5": "Construction",
    "C9+B10": "Construction",
    "B11+C6-C8": "Construction",
    "E1-E2": "Governance / AIDLC tooling",
    "D-PR": "Operation",
    "V1": "Validation",
}


def phase_of(task_id):
    if task_id in PHASE_OF:
        return PHASE_OF[task_id]
    return "Construction"


def write_header(ws, headers, row=1, fill=HEAD):
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=text)
        c.font = BOLD
        c.fill = fill
        c.alignment = WRAP
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def token_rates(region, acus_in_region):
    """Measured tokens per ACU for a telemetry region."""
    return {
        "cache_read": region["cache_read_tokens"] / acus_in_region,
        "cache_write": region["cache_creation_tokens"] / acus_in_region,
        "input": region["input_tokens"] / acus_in_region,
        "output": region["output_tokens"] / acus_in_region,
    }


def sheet_claude_layout(wb, usage, report, tel, rate, rates):
    ws = wb.create_sheet("AIDLC token & ACU")
    ws["A1"] = "DEVIN - Angular 9 -> 20 migration: AIDLC token & ACU consumption"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (
        f"Devin bills ACUs, not tokens. Dollar Value = ACUs x ${rate:.2f}/ACU. "
        "Token / Input / Cache Write / Cache Read / Output are measured context telemetry "
        "(context_growth_update), shown for parity with the Claude sheet - they carry no price."
    )
    ws["A2"].alignment = WRAP
    ws["A3"] = f"Session {report['session_id']}  |  {usage['repository']}  |  PR {usage['pull_request']}"

    headers = ["AIDLC Phases", "Task", "Model", "Token", "Dollar Value",
               "Input", "Cache Write", "Cache Read", "Output", "ACUs", "Basis"]
    write_header(ws, headers, row=5)

    row = 6
    first = row
    for t in usage["tasks"]:
        acus = t["acus"]
        r = rates["construction"]
        ws.cell(row=row, column=1, value=phase_of(t["id"]))
        ws.cell(row=row, column=2, value=f'{t["id"]}  {t["task"]}').alignment = WRAP
        ws.cell(row=row, column=3, value=MODEL_CELL)
        ws.cell(row=row, column=4, value=f"=H{row}+G{row}").number_format = "#,##0"
        ws.cell(row=row, column=5, value=f"=J{row}*{rate}").number_format = '"$"#,##0.00'
        ws.cell(row=row, column=6, value=round(acus * r["input"])).number_format = "#,##0"
        ws.cell(row=row, column=7, value=round(acus * r["cache_write"])).number_format = "#,##0"
        ws.cell(row=row, column=8, value=round(acus * r["cache_read"])).number_format = "#,##0"
        ws.cell(row=row, column=9, value=round(acus * r["output"])).number_format = "#,##0"
        ws.cell(row=row, column=10, value=round(acus, 4))
        ws.cell(row=row, column=11, value="ACUs derived (wall-clock split of measured total); "
                                          "tokens = ACUs x measured construction token rate").alignment = WRAP
        row += 1

    mig_last = row - 1
    ws.cell(row=row, column=2, value="Migration delivery subtotal").font = BOLD
    for col in (4, 6, 7, 8, 9, 10):
        L = get_column_letter(col)
        c = ws.cell(row=row, column=col, value=f"=SUM({L}{first}:{L}{mig_last})")
        c.font = BOLD
        c.number_format = "#,##0" if col != 10 else "0.0000"
    ws.cell(row=row, column=5, value=f"=J{row}*{rate}").number_format = '"$"#,##0.00'
    ws.cell(row=row, column=5).font = BOLD
    ws.cell(row=row, column=11, value="ACU subtotal is MEASURED (20.2524 ACUs, session API)").alignment = WRAP
    row += 2

    rep_first = row
    for e in report["interaction_ledger"]:
        if e["category"] != "Reporting":
            continue
        acus = e["delta_acus"]
        r = rates["reporting"]
        ws.cell(row=row, column=1, value="Reporting (post-delivery)")
        ws.cell(row=row, column=2, value=f'#{e["n"]} {e["paid_for"]}').alignment = WRAP
        ws.cell(row=row, column=3, value=MODEL_CELL)
        ws.cell(row=row, column=4, value=f"=H{row}+G{row}").number_format = "#,##0"
        ws.cell(row=row, column=5, value=f"=J{row}*{rate}").number_format = '"$"#,##0.00'
        ws.cell(row=row, column=6, value=round(acus * r["input"])).number_format = "#,##0"
        ws.cell(row=row, column=7, value=round(acus * r["cache_write"])).number_format = "#,##0"
        ws.cell(row=row, column=8, value=round(acus * r["cache_read"])).number_format = "#,##0"
        ws.cell(row=row, column=9, value=round(acus * r["output"])).number_format = "#,##0"
        ws.cell(row=row, column=10, value=round(acus, 4))
        ws.cell(row=row, column=11, value="ACUs MEASURED (per-interaction checkpoint delta); "
                                          "tokens = ACUs x measured reporting token rate").alignment = WRAP
        row += 1
    rep_last = row - 1
    ws.cell(row=row, column=2, value="Reporting subtotal").font = BOLD
    for col in (4, 6, 7, 8, 9, 10):
        L = get_column_letter(col)
        c = ws.cell(row=row, column=col, value=f"=SUM({L}{rep_first}:{L}{rep_last})")
        c.font = BOLD
        c.number_format = "#,##0" if col != 10 else "0.0000"
    ws.cell(row=row, column=5, value=f"=J{row}*{rate}").number_format = '"$"#,##0.00'
    ws.cell(row=row, column=5).font = BOLD
    row += 2

    ws.cell(row=row, column=2, value="SESSION TOTAL").font = BOLD
    for col in (4, 6, 7, 8, 9, 10):
        L = get_column_letter(col)
        c = ws.cell(row=row, column=col,
                    value=f"=SUM({L}{first}:{L}{mig_last})+SUM({L}{rep_first}:{L}{rep_last})")
        c.font = BOLD
        c.number_format = "#,##0" if col != 10 else "0.0000"
    c = ws.cell(row=row, column=5, value=f"=J{row}*{rate}")
    c.font = BOLD
    c.number_format = '"$"#,##0.00'
    ws.cell(row=row, column=11, value="Session ACU total MEASURED: 34.1175").alignment = WRAP
    for r_ in (row,):
        for col in range(1, 12):
            ws.cell(row=r_, column=col).fill = SUB
    autosize(ws, [22, 52, 26, 14, 13, 12, 13, 14, 12, 10, 46])
    return ws


def sheet_rate_card(wb, tel, rate, rates):
    ws = wb.create_sheet("Rate card & token rates")
    ws["A1"] = "What Devin charges for"
    ws["A1"].font = Font(bold=True, size=12)
    write_header(ws, ["Billable unit", "Price", "Billed?", "Note"], row=3)
    rows = [
        ("ACU (Agent Compute Unit)", f"${rate:.2f} per ACU", "YES",
         "Only billable unit. Enterprise price is set on the order form; $2.25 is the list-price assumption used here."),
        ("Input tokens", "$0.00", "No", "Not billed and not exposed as a billing field by Devin."),
        ("Cache write tokens", "$0.00", "No", "Not billed. Prompt caching shows up only as fewer ACUs."),
        ("Cache read tokens", "$0.00", "No", "Not billed."),
        ("Output tokens", "$0.00", "No", "Not billed."),
    ]
    r = 4
    for a, b, c, d in rows:
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        ws.cell(row=r, column=3, value=c).font = BOLD
        ws.cell(row=r, column=4, value=d).alignment = WRAP
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Measured Devin token rates (tokens per ACU)").font = Font(bold=True, size=12)
    r += 1
    write_header(ws, ["Region", "Telemetry window", "Iterations", "ACUs in window",
                      "Cache read / ACU", "Cache write / ACU", "Input / ACU", "Output / ACU"], row=r)
    r += 1
    for name, key, acus in (("Construction (migration)", "construction_measured", rates["construction_acus"]),
                            ("Reporting (post-delivery Q&A)", "reporting_measured", rates["reporting_acus"])):
        reg = tel["regions"][key]
        k = "construction" if key.startswith("construction") else "reporting"
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=f'{reg["window_utc"][0]} -> {reg["window_utc"][1]} (iterations {reg["iterations"][0]}-{reg["iterations"][1]}, {reg["samples"]} samples)').alignment = WRAP
        ws.cell(row=r, column=3, value=reg["iteration_steps"])
        ws.cell(row=r, column=4, value=round(acus, 4))
        for col, kk in ((5, "cache_read"), (6, "cache_write"), (7, "input"), (8, "output")):
            ws.cell(row=r, column=col, value=round(rates[k][kk])).number_format = "#,##0"
        r += 1
    r += 1
    ws.cell(row=r, column=1, value=(
        "Every token cell in the 'AIDLC token & ACU' sheet = that row's ACUs x the rate above for its region. "
        "The rates themselves are measured, not assumed: they are the context telemetry of the two fully sampled "
        "windows divided by the ACUs consumed in the same windows.")).alignment = WRAP
    autosize(ws, [30, 52, 12, 15, 18, 18, 14, 14])


def sheet_cache(wb, tel):
    ws = wb.create_sheet("Cache efficiency")
    ws["A1"] = "Prompt cache / context reuse (measured)"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = tel["field_mapping"]["cache_read_tokens"]
    ws["A2"].alignment = WRAP
    write_header(ws, ["Window", "Iterations", "Prompt tokens processed", "Cache read (reused prefix)",
                      "Cache write (new tokens)", "Hit rate %", "New tokens / iteration",
                      "Avg context size", "Input tokens", "Output tokens"], row=4)
    r = 5
    for name, key in (("Construction, iterations 1-57", "construction_measured"),
                      ("Reporting, iterations 200-225", "reporting_measured")):
        reg = tel["regions"][key]
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=reg["iteration_steps"])
        for col, k in ((3, "prompt_tokens"), (4, "cache_read_tokens"), (5, "cache_creation_tokens")):
            ws.cell(row=r, column=col, value=reg[k]).number_format = "#,##0"
        ws.cell(row=r, column=6, value=reg["hit_rate_pct"])
        ws.cell(row=r, column=7, value=reg["new_tokens_per_iteration"]).number_format = "#,##0"
        ws.cell(row=r, column=8, value=reg["avg_context_tokens"]).number_format = "#,##0"
        ws.cell(row=r, column=9, value=reg["input_tokens"]).number_format = "#,##0"
        ws.cell(row=r, column=10, value=reg["output_tokens"]).number_format = "#,##0"
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Findings").font = BOLD
    r += 1
    facts = [
        "98.47% of the tokens processed during construction were reused prefix, not new tokens - the stable "
        "prompt prefix (package.json + angular.json + tsconfigs + file tree + conventions) did its job.",
        "Reporting turns write 3,234 new tokens per iteration vs 1,216 during construction: explaining the "
        "migration is more token-dense per iteration than performing it.",
        f'Peak context reached {tel["peak_context_tokens"]:,} tokens; two compaction events '
        "(iterations 197->200 and 221->225) discarded context and reset the cached prefix, which is why "
        "negative deltas are excluded from cache-write.",
        "Devin never charges for any of this: cache efficiency reaches the invoice only as fewer ACUs.",
    ]
    for f in facts:
        ws.cell(row=r, column=1, value=f).alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        r += 1
    autosize(ws, [46, 11, 15, 16, 16, 11, 14, 14, 13, 13])


def sheet_shadow(wb, tel, rate, rates, usage, report):
    ws = wb.create_sheet("If tokens were billed")
    ws["A1"] = "Devin's token volume priced at the Claude rate card (comparison only - Devin bills none of this)"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = ("Rates below are the ones in the customer's Anthropic workbook. This sheet answers the obvious "
                "leadership question: 'how does the ACU bill compare with a token bill for the same work?' "
                "It is a comparison, not an invoice.")
    ws["A2"].alignment = WRAP
    ws["A2"].fill = WARN

    write_header(ws, ["Model", "Input $/M", "Cache write $/M", "Cache read $/M", "Output $/M"], row=4)
    r = 5
    for name, rr in CLAUDE_RATES.items():
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=rr["input"])
        ws.cell(row=r, column=3, value=rr["cache_write"])
        ws.cell(row=r, column=4, value=rr["cache_read"])
        ws.cell(row=r, column=5, value=rr["output"])
        r += 1

    mig_acus = usage["measured"]["migration_acus"]
    rep_acus = sum(e["delta_acus"] for e in report["interaction_ledger"] if e["category"] == "Reporting")
    totals = {k: mig_acus * rates["construction"][k] + rep_acus * rates["reporting"][k]
              for k in ("input", "cache_write", "cache_read", "output")}

    r += 1
    write_header(ws, ["Scope", "Input tokens", "Cache write tokens", "Cache read tokens", "Output tokens",
                      "Token cost @ Opus", "Token cost @ Sonnet 4.6", "Actual Devin bill (ACUs)"], row=r)
    r += 1
    for scope, acus, key in (("Migration delivery", mig_acus, "construction"),
                             ("Post-delivery reporting", rep_acus, "reporting"),
                             ("Session total", None, None)):
        if acus is None:
            vals = totals
            acu_v = mig_acus + rep_acus
        else:
            vals = {k: acus * rates[key][k] for k in ("input", "cache_write", "cache_read", "output")}
            acu_v = acus
        ws.cell(row=r, column=1, value=scope)
        for col, k in ((2, "input"), (3, "cache_write"), (4, "cache_read"), (5, "output")):
            ws.cell(row=r, column=col, value=round(vals[k])).number_format = "#,##0"
        for col, model in ((6, "Opus"), (7, "Sonnet 4.6")):
            rr = CLAUDE_RATES[model]
            cost = (vals["input"] * rr["input"] + vals["cache_write"] * rr["cache_write"]
                    + vals["cache_read"] * rr["cache_read"] + vals["output"] * rr["output"]) / 1e6
            ws.cell(row=r, column=col, value=round(cost, 2)).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=8, value=round(acu_v * rate, 2)).number_format = '"$"#,##0.00'
        r += 1
    r += 1
    ws.cell(row=r, column=1, value=(
        "Read this carefully before quoting it: the token bill above covers model inference only. The ACU bill "
        "covers inference plus the VM, the browser, every npm install and build, planning and the agent loop. "
        "They are not like-for-like, and a low token bill does not mean the same work could be bought for that price.")).alignment = WRAP
    ws.cell(row=r, column=1).fill = WARN
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    autosize(ws, [26, 16, 18, 18, 15, 17, 20, 22])


def sheet_samples(wb, tel):
    ws = wb.create_sheet("Raw telemetry")
    ws["A1"] = tel["source"]
    ws["A1"].alignment = WRAP
    write_header(ws, ["Timestamp (UTC)", "Iteration", "Context tokens", "Context bytes",
                      "Tool-output bytes (cum.)", "Tool-invocation bytes (cum.)",
                      "Delta context tokens vs previous sample"], row=3)
    r = 4
    prev = None
    for s in tel["samples"]:
        ws.cell(row=r, column=1, value=s["ts"][:19] + "Z")
        ws.cell(row=r, column=2, value=s["it"])
        ws.cell(row=r, column=3, value=s["ctx_tok"]).number_format = "#,##0"
        ws.cell(row=r, column=4, value=s["ctx_bytes"]).number_format = "#,##0"
        ws.cell(row=r, column=5, value=s["tool_out_bytes"]).number_format = "#,##0"
        ws.cell(row=r, column=6, value=s["tool_inv_bytes"]).number_format = "#,##0"
        if prev is not None:
            c = ws.cell(row=r, column=7, value=s["ctx_tok"] - prev)
            c.number_format = "#,##0"
            if s["ctx_tok"] < prev:
                c.fill = WARN
        prev = s["ctx_tok"]
        r += 1
    autosize(ws, [22, 11, 15, 14, 22, 24, 30])


def sheet_provenance(wb, usage, report, tel, rate, rates):
    ws = wb.create_sheet("Provenance & formulas")
    write_header(ws, ["Figure", "Value", "How it was obtained", "Basis"], row=1)
    mig = usage["measured"]["migration_acus"]
    rep = sum(e["delta_acus"] for e in report["interaction_ledger"] if e["category"] == "Reporting")
    A = tel["regions"]["construction_measured"]
    B = tel["regions"]["reporting_measured"]
    rows = [
        ("Migration ACUs", round(mig, 4),
         "Devin session API: acus_consumed at the end of migration work minus 0 at session start "
         "(2026-08-28 17:06:16Z -> 17:52:44Z)", "MEASURED"),
        ("Post-delivery reporting ACUs", round(rep, 4),
         "Sum of acu_consumption_at_last_user_interaction deltas for the 7 reporting requests", "MEASURED"),
        ("Session ACUs", round(mig + rep, 4), "Sum of the two rows above", "MEASURED"),
        ("Price per ACU", rate,
         "Devin list-price assumption. Enterprise rate is set on your order form; re-run with --rate to change it",
         "ASSUMPTION"),
        ("Dollar Value column", "ACUs x rate", "Live formula in the workbook, recomputes if you edit the rate",
         "MEASURED x ASSUMPTION"),
        ("Per-task ACUs", "17 rows", "Measured migration total split pro rata by each task's wall-clock duration, "
         "using commit timestamps as boundaries. Devin exposes no per-task ACU counter", "DERIVED"),
        ("Cache read tokens", f'{A["cache_read_tokens"]:,} in the construction window',
         "context_growth_update: the context resident before each iteration, i.e. the prefix a cache serves",
         "MEASURED (proxy for Claude's cache_read_input_tokens)"),
        ("Cache write tokens", f'{A["cache_creation_tokens"]:,} in the construction window',
         "Positive delta of current_context_tokens between iterations", "MEASURED (proxy)"),
        ("Input tokens", f'{A["input_tokens"]:,} in the construction window',
         "Share of new tokens attributable to tool results (delta total_tool_output_bytes / measured bytes-per-token)",
         "MEASURED (proxy)"),
        ("Output tokens", f'{A["output_tokens"]:,} in the construction window',
         "New tokens minus tool-result tokens = model-generated text and tool calls", "MEASURED (proxy)"),
        ("Token rate per ACU", "see 'Rate card & token rates'",
         f'Construction: window tokens / {rates["construction_acus"]:.4f} ACUs consumed in the same window. '
         f'Reporting: window tokens / {rates["reporting_acus"]:.4f} ACUs.', "MEASURED"),
        ("Per-row token cells", "ACUs x token rate",
         "Applied per region. Rows inherit the ACU basis of their row (measured for reporting, derived for tasks)",
         "DERIVED from measured rates"),
        ("Model / agent mode", "Not exposed", "The Devin session API returns no model identity or agent mode for a "
         "completed session, and Devin publishes no per-mode ACU multiplier. Any Ultra-vs-Normal number would be "
         "invented; measure it with a controlled A/B re-run instead", "NOT AVAILABLE"),
        ("Claude rate card", "Opus 5/6.25/0.5/25, Sonnet 4.6 3/2.5/0.3/15 USD per million",
         "Taken from the customer's Anthropic AIDLC workbook, used only on the 'If tokens were billed' sheet",
         "CUSTOMER-SUPPLIED"),
        ("Migration outcome", "PR #35, lint clean, 6/6 unit specs, dev+prod build, 3/3 Playwright e2e",
         "Local command output on branch devin/1787936828-ng9-to-ng20-migration", "MEASURED"),
        ("Browser test run", "Not completed",
         "The persistent testing-agent run was suspended with usage_limit_exceeded; no recording or screenshots exist",
         "NOT COMPLETED"),
    ]
    r = 2
    for a, b, c, d in rows:
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b).alignment = WRAP
        ws.cell(row=r, column=3, value=c).alignment = WRAP
        ws.cell(row=r, column=4, value=d).alignment = WRAP
        r += 1
    autosize(ws, [26, 34, 78, 34])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--rate", type=float, default=2.25)
    args = ap.parse_args()

    usage = json.loads(USAGE.read_text())
    report = json.loads(REPORT.read_text())
    tel = json.loads(TELEMETRY.read_text())

    # ACUs consumed inside each fully sampled telemetry window.
    # Construction window = iterations 1-57 = 17:06:25 -> 17:18:47, i.e. the first
    # 741 s of the 2126 s migration; ACUs allocated by the same wall-clock rule.
    mig_acus = usage["measured"]["migration_acus"]
    total_secs = sum(t["duration_seconds"] for t in usage["tasks"])
    construction_acus = mig_acus * 741 / total_secs
    # Reporting window = iterations 200-225 = the 09-01 11:0x interaction, whose
    # measured ACU delta is the last reporting checkpoint.
    reporting_acus = report["interaction_ledger"][-1]["delta_acus"]

    rates = {
        "construction": token_rates(tel["regions"]["construction_measured"], construction_acus),
        "reporting": token_rates(tel["regions"]["reporting_measured"], reporting_acus),
        "construction_acus": construction_acus,
        "reporting_acus": reporting_acus,
    }

    wb = Workbook()
    wb.remove(wb.active)
    sheet_claude_layout(wb, usage, report, tel, args.rate, rates)
    sheet_rate_card(wb, tel, args.rate, rates)
    sheet_cache(wb, tel)
    sheet_shadow(wb, tel, args.rate, rates, usage, report)
    sheet_samples(wb, tel)
    sheet_provenance(wb, usage, report, tel, args.rate, rates)
    wb.save(OUT)
    print(f"wrote {OUT} at ${args.rate:.2f}/ACU")
    print(f"construction window ACUs {construction_acus:.4f}, reporting window ACUs {reporting_acus:.4f}")


if __name__ == "__main__":
    main()
